#!/usr/bin/env python3
"""
Klein Finance - Monthly Updater v1.0
Auto-updates itself from GitHub on each run.
"""

import os, sys, re, shutil, json, urllib.request
from pathlib import Path
from datetime import datetime

VERSION    = "2.0"
UPDATE_URL = "https://gist.githubusercontent.com/claude-klein-finance/raw/update.py"
BASE       = Path(__file__).parent
MONTHLY    = BASE / "monthly"
BACKUPS    = BASE / "backups"
# Load Anthropic API key from local file (never stored in GitHub)
ANTHROPIC_KEY = ""
_key_file = BASE / "api_key.txt"
if _key_file.exists():
    ANTHROPIC_KEY = _key_file.read_text(encoding="utf-8").strip()

EXCEL      = BASE / "\u05de\u05d0\u05d6\u05df_\u05e7\u05dc\u05d9\u05d9\u05df.xlsm"

missing = []
try:    import pandas as pd
except: missing.append("pandas")
try:    import openpyxl
except: missing.append("openpyxl")
if missing:
    print(f"Run: python -m pip install {' '.join(missing)}")
    sys.exit(1)

from openpyxl import load_workbook

G="\033[32m"; Y="\033[33m"; R="\033[31m"; C="\033[36;1m"; X="\033[0m"
def ok(t):   print(f"  {G}\u2713{X} {t}")
def warn(t): print(f"  {Y}\u26a0{X} {t}")
def err(t):  print(f"  {R}\u2717{X} {t}")
def hdr(t):  print(f"\n{C}\u2500\u2500 {t} \u2500\u2500{X}")

def num(val):
    if val is None: return 0.0
    try: return float(str(val).replace(",","").replace("\u20aa","").replace("$","").replace("%","").replace(" ","").strip())
    except: return 0.0

def read_html_xls(path):
    """Read Leumi HTML files disguised as .xls"""
    for enc in ["windows-1255", "utf-8", "iso-8859-8"]:
        try:
            tables = pd.read_html(str(path), encoding=enc)
            if tables: return tables[0]
        except: pass
    return None

def detect_type(path):
    name = path.name
    if "\u05e2\u05d5\u05e9" in name or "\u05dc\u05d0\u05d5\u05de\u05d9 \u05e9\u05e2" in name or "\u05e2\u05d5\u05e9 \u05dc\u05d0\u05d5\u05de\u05d9" in name:
        return "bank"
    if "\u05d4\u05ea\u05de\u05d5\u05e0\u05d4 \u05d4\u05de\u05dc\u05d0\u05d4" in name:
        return "pension_liat" if "(11)" in name else "pension_dror"
    if "\u05d0\u05d7\u05d6\u05e7\u05d5\u05ea" in name:
        return "invest"
    if "\u05e8\u05d9\u05db\u05d5\u05d6 \u05d9\u05ea\u05e8\u05d5\u05ea" in name or ("\u05e8\u05d9\u05db\u05d5\u05d6" in name and "\u05d9\u05ea\u05e8\u05d5\u05ea" in name):
        return "balance"
    if "5647" in name or "\u05d0\u05d9\u05e9\u05e8\u05d0\u05db\u05e8\u05d8" in name.lower():
        return "isracard"
    if "transaction-details" in name.lower():
        try:
            df = pd.read_excel(path, header=None, nrows=10, engine="openpyxl")
            text = " ".join(str(v) for row in df.values for v in row if str(v) != "nan")
            return "foreign" if '\u05d7\u05d5"\u05dc' in text or '\u05de\u05d8"\u05d7' in text else "credit"
        except: return "credit"
    try:
        engine = "xlrd" if str(path).endswith(".xls") else "openpyxl"
        df = pd.read_excel(path, header=None, nrows=6, engine=engine)
        text = " ".join(str(v) for row in df.values for v in row if str(v) != "nan")
        if "\u05d9\u05ea\u05e8\u05d4 \u05de\u05e6\u05d8\u05d1\u05e8\u05ea" in text: return "bank"
        if "\u05e9\u05dd \u05de\u05d5\u05e6\u05e8" in text and "\u05e4\u05d5\u05dc\u05d9\u05e1\u05d4" in text:
            df2 = pd.read_excel(path, header=None, engine=engine)
            rows = sum(1 for _,r in df2.iterrows() if any(str(v)!="nan" for v in r))
            return "pension_liat" if rows <= 8 else "pension_dror"
        if "\u05e9\u05dd \u05d4\u05e0\u05d9\u05d9\u05e8" in text or "\u05de\u05d1\u05d8 \u05d0\u05d9\u05e9\u05d9" in text: return "invest"
        if "\u05ea\u05d0\u05e8\u05d9\u05da \u05e8\u05db\u05d9\u05e9\u05d4" in text and "\u05e9\u05dd \u05d1\u05d9\u05ea \u05e2\u05e1\u05e7" in text: return "isracard"
        if "\u05ea\u05d0\u05e8\u05d9\u05da \u05e2\u05e1\u05e7\u05d4" in text and "\u05e1\u05d5\u05d2 \u05e2\u05e1\u05e7\u05d4" in text:
            return "foreign" if "\u05de\u05d8\u05d1\u05e2 \u05e2\u05e1\u05e7\u05d4 \u05de\u05e7\u05d5\u05e8\u05d9" in text else "credit"
    except: pass
    return None

def find_files():
    hdr("Scanning monthly folder")
    MONTHLY.mkdir(exist_ok=True)
    excel_files = list(MONTHLY.glob("*.xls")) + list(MONTHLY.glob("*.xlsx")) + list(MONTHLY.glob("*.xlsm"))
    image_files = list(MONTHLY.glob("*.png")) + list(MONTHLY.glob("*.jpg")) + list(MONTHLY.glob("*.jpeg"))
    found = {k: None for k in ["bank","credit","foreign","isracard","pension_dror","pension_liat","invest","balance","rsu_image"]}
    for f in excel_files:
        ft = detect_type(f)
        if ft and found[ft] is None:
            found[ft] = f; ok(f"Found {ft}: {f.name}")
        elif ft is None:
            warn(f"Could not identify: {f.name}")
    for f in image_files:
        found["rsu_image"] = f; ok(f"Found RSU image: {f.name}")
    return found

def parse_bank(path):
    engine = "xlrd" if str(path).endswith(".xls") else "openpyxl"
    df = pd.read_excel(path, header=None, engine=engine)
    balance = 0
    try:
        b = str(df.iloc[2,0]).replace("\u20aa","").replace(",","").replace(" ","")
        balance = float(b)
    except: pass
    txs = []
    for i,row in df.iterrows():
        if i < 2: continue
        row = list(row)
        d = num(row[6]) if len(row)>6 else 0
        c = num(row[7]) if len(row)>7 else 0
        if d==0 and c==0: continue
        txs.append({"debit":d,"credit":c})
    return {"balance":balance,"income":sum(t["credit"] for t in txs),"expense":sum(t["debit"] for t in txs),"raw_df":df}

def parse_credit(path):
    engine = "xlrd" if str(path).endswith(".xls") else "openpyxl"
    df = pd.read_excel(path, header=None, engine=engine)
    txs=[]; hf=False
    for i,row in df.iterrows():
        row=list(row)
        if "\u05ea\u05d0\u05e8\u05d9\u05da \u05e2\u05e1\u05e7\u05d4" in str(row[0]) or "\u05ea\u05d0\u05e8\u05d9\u05da \u05e8\u05db\u05d9\u05e9\u05d4" in str(row[0]):
            hf=True; continue
        if not hf or not row[0] or str(row[0])=="nan": continue
        a = num(row[5]) if len(row)>5 else num(row[2])
        if a==0: continue
        txs.append({"category":str(row[2] or "\u05e9\u05d5\u05e0\u05d5\u05ea"),"amount":a})
    bc={}
    for t in txs: bc[t["category"]]=bc.get(t["category"],0)+t["amount"]
    return {"transactions":txs,"by_category":bc,"total":sum(bc.values()),"raw_df":df}

def parse_isracard(path):
    engine = "xlrd" if str(path).endswith(".xls") else "openpyxl"
    df = pd.read_excel(path, header=None, engine=engine)
    txs=[]
    for i,row in df.iterrows():
        if i==0: continue
        row=list(row)
        if not row[0] or str(row[0])=="nan": continue
        a=num(row[2])
        if a==0: continue
        cat=str(row[8]) if len(row)>8 and str(row[8])!="nan" else "\u05e9\u05d5\u05e0\u05d5\u05ea"
        txs.append({"category":cat,"amount":a})
    bc={}
    for t in txs: bc[t["category"]]=bc.get(t["category"],0)+t["amount"]
    return {"transactions":txs,"by_category":bc,"total":sum(bc.values()),"raw_df":df}

def parse_pension(path):
    df = None
    for engine in ["xlrd", "openpyxl"]:
        try:
            df = pd.read_excel(path, header=None, engine=engine)
            break
        except: pass
    if df is None:
        df = read_html_xls(path)
    if df is None:
        return {}
    products = []
    for i, row in df.iterrows():
        if i == 0: continue
        row = list(row)
        if not row[0] or str(row[0]) == "nan": continue
        product_name = str(row[0])
        t = num(row[4]) if len(row) > 4 else 0
        if t == 0:
            for v in row[2:8]:
                candidate = num(str(v))
                if 1000 < candidate < 10000000:
                    t = candidate; break
        if t == 0: continue
        products.append({"product": product_name, "total": t,
            "fee_deposit": num(row[11]) if len(row) > 11 else 0,
            "fee_accum":   num(row[12]) if len(row) > 12 else 0,
            "return_ytd":  num(row[13]) if len(row) > 13 else 0})
    pension_kw   = ["\u05e4\u05e0\u05e1\u05d9\u05d4"]
    provident_kw = ["\u05d4\u05e9\u05ea\u05dc\u05de\u05d5\u05ea", "\u05e7\u05e8\u05df \u05d4\u05e9\u05ea\u05dc\u05de\u05d5\u05ea"]
    pension_total   = sum(p["total"] for p in products if any(k in p["product"] for k in pension_kw))
    provident_total = sum(p["total"] for p in products if any(k in p["product"] for k in provident_kw))
    if pension_total == 0 and provident_total == 0:
        pension_total = sum(p["total"] for p in products)
    return {"pension": pension_total, "provident": provident_total, "products": products, "raw_df": df}

def parse_invest(path):
    # Try HTML first (Leumi)
    df = read_html_xls(path)
    if df is not None:
        total = 0
        for _, row in df.iterrows():
            for val in row:
                v = num(str(val))
                if 500000 < v < 20000000:
                    total = v; break
            if total: break
        return {"total": total, "raw_df": df}
    try:
        engine = "xlrd" if str(path).endswith(".xls") else "openpyxl"
        df = pd.read_excel(path, header=None, engine=engine)
        total = 0
        try: total = num(str(df.iloc[2,3]).replace(",",""))
        except: pass
        return {"total": total, "raw_df": df}
    except: return {}

def parse_rsu_image(path):
    """Read RSU numbers from screenshot via Claude API."""
    try:
        import base64, json as _j
        with open(path,"rb") as f: b64=base64.b64encode(f.read()).decode()
        mime="image/png" if str(path).lower().endswith(".png") else "image/jpeg"
        payload={"model":"claude-sonnet-4-20250514","max_tokens":100,
                 "messages":[{"role":"user","content":[
                     {"type":"image","source":{"type":"base64","media_type":mime,"data":b64}},
                     {"type":"text","text":'From this RSU equity overview screenshot extract two numbers. Reply JSON only, no other text: {"unvested": 187148, "available": 170600}'}
                 ]}]}
        req=urllib.request.Request("https://api.anthropic.com/v1/messages",
            data=_j.dumps(payload).encode(),
            headers={"Content-Type":"application/json","anthropic-version":"2023-06-01","x-api-key":ANTHROPIC_KEY})
        with urllib.request.urlopen(req,timeout=20) as r:
            res=_j.loads(r.read())
            text=res["content"][0]["text"].strip()
            m=re.search(r'\{[^}]+\}',text)
            if m: return _j.loads(m.group())
    except: pass
    warn("RSU image could not be read. Enter manually:")
    try:
        u=float(input("    Unvested ($): ").replace(",","").replace("$","").strip())
        a=float(input("    Available ($): ").replace(",","").replace("$","").strip())
        return {"unvested":u,"available":a}
    except: return {"unvested":0,"available":0}

def parse_all(found):
    hdr("Parsing files")
    data={}
    parsers={"bank":parse_bank,"credit":parse_credit,"foreign":parse_credit,
             "isracard":parse_isracard,"pension_dror":parse_pension,"pension_liat":parse_pension,
             "invest":parse_invest,"balance":parse_invest,"rsu_image":parse_rsu_image}
    for key,path in found.items():
        if path is None: data[key]={}; warn(f"Missing: {key}"); continue
        try: data[key]=parsers[key](path); ok(f"Parsed: {key}")
        except Exception as e: warn(f"Error ({key}): {e}"); data[key]={}
    return data

def update_mislaka_sheet(wb, sheet_name, pension_data):
    """Update only data rows in pension sheet, preserving structure."""
    if sheet_name not in wb.sheetnames or not pension_data:
        return
    products = pension_data.get("products", [])
    if not products:
        return
    ws = wb[sheet_name]
    # Write product data starting from row 2
    for i, product in enumerate(products, 2):
        ws.cell(row=i, column=1).value = product.get("product", "")
        ws.cell(row=i, column=5).value = product.get("total", 0)
        ws.cell(row=i, column=12).value = product.get("fee_deposit", 0)
        ws.cell(row=i, column=13).value = product.get("fee_accum", 0)
        ws.cell(row=i, column=14).value = product.get("return_ytd", 0)

def update_dashboard(wb, data):
    if "\u05d3\u05e9\u05d1\u05d5\u05e8\u05d3" not in wb.sheetnames: return
    ws=wb["\u05d3\u05e9\u05d1\u05d5\u05e8\u05d3"]

    dror_p = data.get("pension_dror",{}).get("pension",0)
    liat_p = data.get("pension_liat",{}).get("pension",0)
    dror_h = data.get("pension_dror",{}).get("provident",0)
    liat_h = data.get("pension_liat",{}).get("provident",0)
    print(f"  DEBUG pension: dror_p={dror_p}, liat_p={liat_p}, dror_h={dror_h}, liat_h={liat_h}")
    invest = data.get("invest",{}).get("total",0)
    bank   = data.get("bank",{}).get("balance",0)
    rsu    = data.get("rsu_image",{})

    updates=[]
    def set_cell(row, col, val, label):
        if val and val > 0:
            ws.cell(row=row, column=col).value = val
            updates.append(f"{label}: {val:,.0f}")

    set_cell(10, 4, dror_p, "\u05e4\u05e0\u05e1\u05d9\u05d4 \u05d3\u05e8\u05d5\u05e8")
    set_cell(11, 4, liat_p, "\u05e4\u05e0\u05e1\u05d9\u05d4 \u05dc\u05d9\u05d0\u05ea")
    set_cell(12, 4, dror_h, "\u05d4\u05e9\u05ea\u05dc\u05de\u05d5\u05ea \u05d3\u05e8\u05d5\u05e8")
    set_cell(13, 4, liat_h, "\u05d4\u05e9\u05ea\u05dc\u05de\u05d5\u05ea \u05dc\u05d9\u05d0\u05ea")
    set_cell(14, 4, invest, "\u05ea\u05d9\u05e7 \u05d4\u05e9\u05e7\u05e2\u05d5\u05ea")
    set_cell(18, 4, bank,   "\u05e2\u05d5\"\u05e9")

    ws.cell(row=16,column=5).value=rsu.get("available",0)
        updates.append(f"RSU Available: ${rsu['available']:,.0f}")
    ws.cell(row=17,column=5).value=rsu.get("unvested",0)
        updates.append(f"RSU Unvested: ${rsu['unvested']:,.0f}")

    ws.cell(row=2,column=1).value=f"\u05e2\u05d3\u05db\u05d5\u05df \u05d0\u05d7\u05e8\u05d5\u05df: {datetime.now().strftime('%d/%m/%Y')}"
    
    # Also update ÃÂÃÂÃÂÃÂ¡ÃÂÃÂÃÂÃÂ§ÃÂÃÂ sheets with pension data
    update_mislaka_sheet(wb, "\u05d3\u05e8\u05d5\u05e8 - \u05de\u05e1\u05dc\u05e7\u05d4", data.get("pension_dror", {}))
    update_mislaka_sheet(wb, "\u05dc\u05d9\u05d0\u05ea - \u05de\u05e1\u05dc\u05e7\u05d4", data.get("pension_liat", {}))
    for u in updates: ok(f"  {u}")

def replace_sheet_data(wb, sheet_name, df):
    """Only write cell values ÃÂ¢ÃÂÃÂ never touch formatting or structure."""
    if df is None or sheet_name not in wb.sheetnames: return
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row: cell.value = None
    for r, row in enumerate(df.values, 1):
        for c, val in enumerate(row, 1):
            if val is not None and str(val) != "nan":
                try: ws.cell(row=r, column=c).value = val
                except: pass
    ok(f"  Updated: {sheet_name}")

def print_summary(data, found):
    month=datetime.now().strftime("%B %Y")
    print(f"\n{'='*50}\n  Klein Finance \u2014 {month}\n{'='*50}")
    labels={"bank":"\u05e2\u05d5\u05e9","credit":"\u05de\u05e7\u05e1 \u05d0\u05e8\u05e5",
            "foreign":"\u05de\u05e7\u05e1 \u05d7\u05d5\u05dc","isracard":"\u05d0\u05d9\u05e9\u05e8\u05d0\u05db\u05e8\u05d8",
            "pension_dror":"\u05de\u05e1\u05dc\u05e7\u05d4 \u05d3\u05e8\u05d5\u05e8",
            "pension_liat":"\u05de\u05e1\u05dc\u05e7\u05d4 \u05dc\u05d9\u05d0\u05ea",
            "invest":"\u05ea\u05d9\u05e7 \u05d4\u05e9\u05e7\u05e2\u05d5\u05ea",
            "balance":"\u05e8\u05d9\u05db\u05d5\u05d6 \u05d9\u05ea\u05e8\u05d5\u05ea",
            "rsu_image":"RSU"}
    print("\n  Files:")
    for k,l in labels.items():
        sym = f"{G}\u2713{X}" if found.get(k) else f"{Y}\u26a0{X}"
        print(f"    {sym} {l}")
    cc=data.get("credit",{}).get("total",0)+data.get("foreign",{}).get("total",0)
    ic=data.get("isracard",{}).get("total",0)
    if cc+ic: print(f"\n  Total spending: \u20aa{cc+ic:,.0f}")
    print(f"\n  {G}\u2713 Excel updated{X}\n{'='*50}\n")

def main():
    print(f"\n{C}  Klein Family Finance \u2014 Monthly Update v{VERSION}{X}")
    print(f"  {datetime.now().strftime('%d %B %Y, %H:%M')}\n")

    if not EXCEL.exists():
        err(f"Excel not found: {EXCEL}")
        sys.exit(1)

    found = find_files()
    data  = parse_all(found)

    hdr("Backup")
    BACKUPS.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d")
    shutil.copy2(EXCEL, BACKUPS / f"\u05de\u05d0\u05d6\u05df_\u05e7\u05dc\u05d9\u05d9\u05df_{stamp}.xlsm")
    ok("Backup created")

    hdr("Updating Excel")
    wb = load_workbook(EXCEL, keep_vba=True)
    update_dashboard(wb, data)

    # Debug: check what pension data looks like
    pd_data = data.get("pension_dror", {})
    print(f"  DEBUG dror pension dict: pension={pd_data.get('pension',0)}, provident={pd_data.get('provident',0)}")
    products = pd_data.get("products", [])
    for p in products[:5]:
        print(f"    product: {p['product'][:30]} | total: {p['total']}")

    # Only replace transaction sheets - NOT pension/mislaka sheets
    # Pension sheets have formulas that dashboard reads from
    sheet_map = {
        "\u05e2\u05d5\u05e9":                     "bank",
        "\u05e2\u05e1\u05e7\u05d0\u05d5\u05ea \u05d1\u05de\u05d5\u05e2\u05d3 \u05d4\u05d7\u05d9\u05d5\u05d1": "credit",
        '\u05e2\u05e1\u05e7\u05d0\u05d5\u05ea \u05d7\u05d5"\u05dc \u05d5\u05de\u05d8"\u05d7':              "foreign",
        "\u05d0\u05d9\u05e9\u05e8\u05d0\u05db\u05e8\u05d8":              "isracard",
    }
    for sheet, key in sheet_map.items():
        replace_sheet_data(wb, sheet, data.get(key,{}).get("raw_df"))

    wb.save(EXCEL)
    ok("Excel saved")
    print_summary(data, found)

    hdr("Opening Excel")
    os.startfile(str(EXCEL))
    ok("Done!")

if __name__ == "__main__":
    main()
