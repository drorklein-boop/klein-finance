# Klein Finance - Dashboard Generator v3.1.1
import sys, json, warnings, webbrowser, datetime, urllib.request
warnings.filterwarnings('ignore')
from pathlib import Path

BASE = Path(__file__).parent
TEMPLATE_URL = 'https://raw.githubusercontent.com/drorklein-boop/klein-finance/main/dashboard_template.html'
TEMPLATE_LOCAL = BASE / 'dashboard_template.html'

def get_template():
    try:
        url = TEMPLATE_URL + '?t=' + str(int(__import__('time').time()))
        req = urllib.request.Request(url, headers={'Cache-Control': 'no-cache'})
        with urllib.request.urlopen(req, timeout=15) as r:
            content = r.read().decode('utf-8')
            TEMPLATE_LOCAL.write_text(content, encoding='utf-8')
            return content
    except Exception as e:
        print(f'  Warning: could not download template ({e}), using local copy')
        if TEMPLATE_LOCAL.exists():
            return TEMPLATE_LOCAL.read_text(encoding='utf-8')
        raise RuntimeError('No template available')

def ils(n):
    try: return f'₪{abs(float(n)):,.0f}'
    except: return '₪0'

def ils_signed(n):
    try:
        n = float(n)
        return ('+₪' if n >= 0 else '−₪') + f'{abs(n):,.0f}'
    except: return '₪0'

def pct(n, decimals=1):
    try: return f'{float(n)*100:.{decimals}f}%'
    except: return '0%'

def usd(n):
    try: return f'${float(n):,.0f}'
    except: return '$0'

def k(n):
    try: return f'₪{float(n)/1000:,.0f}K'
    except: return '₪0K'

def read_workbook():
    import xlwings as xw
    import openpyxl
    xlsm = next((f for f in BASE.glob('*.xlsm') if not f.name.startswith('~')), None)
    if not xlsm:
        print('ERROR: No .xlsm file found'); return None
    # Use xlwings for formula cells (live values), openpyxl for data-only sheets
    xw_app = xw.apps.active
    if not xw_app:
        print('  ERROR: Excel must be open to generate dashboard')
        return None
    xw_wb = next((b for b in xw_app.books if b.name.lower().endswith('.xlsm')), None)
    if not xw_wb:
        print('  ERROR: No .xlsm workbook open in Excel')
        return None

    # Force full recalculation before reading any values
    xw_wb.app.calculation = 'automatic'
    xw_wb.app.calculate()
    import time as _time; _time.sleep(1)  # wait for calculation to finish
    print(f'  K24 bank value: {xw_wb.sheets["דוח חודשי"].range((24,11)).value}')

    ws_xw = xw_wb.sheets['דוח חודשי']

    def r(row_1based, col_1based):
        val = ws_xw.range((row_1based, col_1based)).value
        return val

    # Also open with openpyxl for sheets that don't have formulas
    wb = openpyxl.load_workbook(xlsm, read_only=True, data_only=True)
    rsu  = wb['ALIGN RSU']

    d = {}
    d['period']          = r(3, 4) or datetime.date.today().strftime('%B %Y')
    d['report_date']     = str(r(3, 2) or datetime.date.today().strftime('%d.%m.%Y'))
    d['income']          = float(r(6, 2) or 0)   # B6
    d['expenses']        = float(r(7, 2) or 0)   # B7
    d['surplus']         = float(r(8, 2) or 0)   # B8
    d['savings_rate']    = float(r(9, 2) or 0)   # B9
    d['portfolio_val']   = float(r(25, 2) or 0)  # B25
    d['kaspiyot']        = float(r(25, 4) or 0)  # D25
    d['portfolio_chg']   = float(r(26, 2) or 0)  # B26
    d['cumulative_gain'] = float(r(27, 2) or 0)  # B27
    d['return_pct']      = float(r(27, 4) or 0)  # D27
    d['monthly_return']  = float(r(29, 4) or 0)  # D29
    d['target_dist']     = float(r(31, 2) or 0)  # B31
    d['invest_surplus']  = float(r(29, 11) or 0) # K29
    d['bank_free']       = float(r(27, 11) or 0) # K27
    d['bank']            = float(r(24, 11) or 0) # K24
    d['exp_credit']      = abs(float(r(25, 11) or 0)) + abs(float(r(26, 11) or 0))  # K25+K26
    d['net_worth']       = float(r(32, 11) or 0) # K32
    d['pension_total']   = float(r(13, 11) or 0) # K13
    d['hishtalmut_total']= float(r(14, 11) or 0) # K14
    d['pension_dror']    = float(r(17, 11) or 0) # K17
    d['pension_liat']    = float(r(18, 11) or 0) # K18
    d['hishtalmut_dror'] = float(r(19, 11) or 0) # K19
    d['hishtalmut_liat'] = float(r(20, 11) or 0) # K20
    d['pension_monthly'] = float(r(21, 11) or 0) # K21
    d['pension_dror_chg']= float(r(17, 12) or 0) # L17
    d['hishtalmut_dror_f1'] = float(r(19, 11) or 0) * 0.47
    d['hishtalmut_dror_f2'] = float(r(19, 11) or 0) * 0.53

    d['rsu_vested']      = float(rsu['H13'].value or 0)
    d['rsu_unvested']    = float(rsu['H14'].value or 0)

    d['pension_dror_chg']  = float(r(17, 12) or 0)  # L17
    d['pension_monthly']   = float(r(21, 11) or 0)  # K21
    d['hishtalmut_dror_f1']= d['hishtalmut_dror'] * 0.47
    d['hishtalmut_dror_f2']= d['hishtalmut_dror'] * 0.53
    # Read actual Migdal Makefet fund split from מסלקה sheet
    try:
        ws_dror = wb['דרור - מסלקה']
        dror_rows = list(ws_dror.iter_rows(values_only=True))
        makefet_vals = [float(r[4]) for r in dror_rows
                        if r[1] and 'מגדל מקפת' in str(r[1]) and r[4] and isinstance(r[4], (int,float)) and float(r[4]) > 1000]
        if len(makefet_vals) >= 2:
            d['pension_dror_f1'] = makefet_vals[0]
            d['pension_dror_f2'] = makefet_vals[1]
        elif len(makefet_vals) == 1:
            d['pension_dror_f1'] = makefet_vals[0]
            d['pension_dror_f2'] = d['pension_dror'] - makefet_vals[0]
        else:
            d['pension_dror_f1'] = d['pension_dror'] * 0.5
            d['pension_dror_f2'] = d['pension_dror'] * 0.5
    except:
        d['pension_dror_f1'] = d['pension_dror'] * 0.5
        d['pension_dror_f2'] = d['pension_dror'] * 0.5

    # Mortgage from balance sheet
    d['mortgage'] = 244691.90
    try:
        ws_bal = wb['ריכוז יתרות לאומי']
        for row in ws_bal.iter_rows(values_only=True):
            if row[0] and 'משכנתא' in str(row[0]) and row[2] and isinstance(row[2], (int, float)):
                d['mortgage'] = abs(float(row[2]))
                break
    except: pass

    # Top 5
    d['top5'] = []
    for row_i in range(13, 18):  # rows 13-17
        rank = ws_xw.range((row_i, 1)).value
        name = ws_xw.range((row_i, 2)).value
        amt  = ws_xw.range((row_i, 3)).value
        src  = ws_xw.range((row_i, 4)).value
        if rank and isinstance(rank, (int,float)) and name and amt:
            d['top5'].append({'name': str(name), 'amount': float(amt), 'source': str(src or '')})

    # Top 3 cats
    d['top3'] = []
    for row_i in range(20, 23):  # rows 20-22
        name = ws_xw.range((row_i, 1)).value
        amt  = ws_xw.range((row_i, 2)).value
        pct_ = ws_xw.range((row_i, 3)).value
        prev = ws_xw.range((row_i, 4)).value
        if name and amt and isinstance(amt, (int,float)):
            d['top3'].append({'name': str(name), 'amount': float(amt),
                              'pct': float(pct_ or 0), 'prev_pct': float(prev or 0)})

    # Chart historical data — read via xlwings for live formula values
    try:
        ws_tz = xw_wb.sheets['ניתוח תזרים']
        def tz(row_1, col_1):
            v = ws_tz.range((row_1, col_1)).value
            return int(float(v)) if isinstance(v, (int,float)) and v else 0
        def tz_str(row_1, col_1):
            v = ws_tz.range((row_1, col_1)).value
            return str(v) if v else ''
        data_cols = [c for c in range(2, 20)
                     if tz(47, c) > 0 and 'סה' not in tz_str(46, c) and tz_str(46, c)]
        d['chart_months']   = [tz_str(46, c) for c in data_cols]
        d['chart_income']   = [tz(47, c) for c in data_cols]
        d['chart_expenses'] = [tz(48, c) for c in data_cols]
        d['chart_invest']   = [tz(49, c) for c in data_cols]
        d['chart_surplus']     = [tz(52, c) for c in data_cols]
        d['chart_align']       = [tz(56, c) for c in data_cols]
        d['chart_insightech']  = [tz(57, c) for c in data_cols]
        print(f'  Chart: {len(data_cols)} months: {d["chart_months"]}')
    except Exception as e:
        print(f'  Chart data read failed: {e}')
        d['chart_months'] = []
        d['chart_income'] = d['chart_expenses'] = d['chart_invest'] = d['chart_surplus'] = []

    # Expense breakdown — from ניתוח תזרים rows 66-70, current month col
    try:
        # Find current month column — same logic as chart (col with income > 0)
        cur_col = data_cols[-1] if data_cols else None
        if cur_col:
            exp_credit   = abs(tz(66, cur_col))
            exp_cash     = abs(tz(68, cur_col))
            exp_mortgage = abs(tz(67, cur_col))
            exp_other    = abs(tz(70, cur_col))
            exp_total    = exp_credit + exp_cash + exp_mortgage + exp_other or 1
            def pct_str(v): return f'{v/exp_total*100:.1f}'
            d['exp_credit_pct']   = pct_str(exp_credit)
            d['exp_cash']         = f'₪{exp_cash:,.0f}'
            d['exp_cash_pct']     = pct_str(exp_cash)
            d['exp_mortgage']     = f'₪{exp_mortgage:,.0f}'
            d['exp_mortgage_pct'] = pct_str(exp_mortgage)
        else:
            d['exp_credit_pct'] = d['exp_cash_pct'] = d['exp_mortgage_pct'] = '0'
            d['exp_cash'] = d['exp_mortgage'] = '₪0'
    except Exception as e:
        print(f'  Expense breakdown failed: {e}')
        d['exp_credit_pct'] = d['exp_cash_pct'] = d['exp_mortgage_pct'] = '0'
        d['exp_cash'] = d['exp_mortgage'] = '₪0'

    # Category computation — read קטגוריה column directly from CAL sheet
    try:
        import datetime as dt
        now = dt.date.today()
        months_back = []
        for i in range(2, -1, -1):
            m = (now.month - i - 1) % 12 + 1
            y = now.year if (now.month - i) > 0 else now.year - 1
            months_back.append((y, m))

        month_names_he = {1:'ינואר',2:'פברואר',3:'מרץ',4:'אפריל',5:'מאי',6:'יוני',
                          7:'יולי',8:'אוגוסט',9:'ספטמבר',10:'אוקטובר',11:'נובמבר',12:'דצמבר'}

        cat_totals = {}  # {cat: [m0, m1, m2]}

        ws_cal = wb['עסקאות במועד החיוב']
        rows_cal = list(ws_cal.iter_rows(values_only=True))
        # Header is row 4 (index 3)
        hdr_idx = next((i for i,r in enumerate(rows_cal)
                        if r[0] and 'תאריך' in str(r[0]) and r[1] and 'שם' in str(r[1])), 3)
        # col 0=date, 1=merchant, 2=category, 5=amount
        for row in rows_cal[hdr_idx+1:]:
            try:
                date_val = row[0]
                if not date_val: continue
                if hasattr(date_val, 'year') and date_val.year > 2000:
                    ry, rm = date_val.year, date_val.month
                else:
                    continue
                month_idx = next((i for i,(y,m) in enumerate(months_back) if y==ry and m==rm), None)
                if month_idx is None: continue
                cat = str(row[2] or 'שונות').strip()
                amount = abs(float(row[5] or 0))
                if cat not in cat_totals: cat_totals[cat] = [0, 0, 0]
                cat_totals[cat][month_idx] += amount
            except: continue

        cats_with_data = sorted(
            [(cat, v) for cat, v in cat_totals.items() if sum(v) > 0],
            key=lambda x: -sum(x[1])
        )
        cats_json = ',\n  '.join(
            f"{{name:\'{cat}\', vals:[{int(v[0])},{int(v[1])},{int(v[2])}]}}"
            for cat, v in cats_with_data[:8]
        )
        cat_months = ','.join(f"'{month_names_he[m]}'" for y,m in months_back)
        d['cats_json']  = cats_json
        d['cat_months'] = cat_months
        d['cat_range']  = f"{month_names_he[months_back[0][1]]} עד {month_names_he[months_back[2][1]]}"
        # Build top3 from computed categories (current month = last index)
        top3_from_cats = cats_with_data[:3]
        total_cur = sum(v[2] for _, v in top3_from_cats) or 1
        d['top3'] = [{'name': cat, 'amount': v[2], 'pct': v[2]/total_cur, 'prev_pct': v[0]/total_cur}
                     for cat, v in top3_from_cats]
        # Category change stats
        # Compare current month (index 2) vs previous month (index 1)
        prev_month_name = month_names_he[months_back[1][1]]
        d['cat_range_start'] = prev_month_name
        total_prev = sum(v[1] for _, v in cats_with_data)
        total_cur2 = sum(v[2] for _, v in cats_with_data)
        if total_prev > 0:
            chg = (total_cur2 - total_prev) / total_prev
            d['cat_total_chg'] = f"{'▲' if chg>=0 else '▼'}{abs(chg)*100:.1f}%"
            d['cat_total_chg_color'] = 'c-red' if chg >= 0 else 'c-green'
        else:
            d['cat_total_chg'] = '—'
            d['cat_total_chg_color'] = 'c-muted'
        # Per-category changes vs previous month
        changes = []
        for cat, v in top3_from_cats:
            if v[1] > 0:
                c = (v[2] - v[1]) / v[1]
                arrow = '▲' if c >= 0 else '▼'
                clr = 'c-red' if c >= 0 else 'c-green'
                changes.append(f'{cat} <span class="{clr}">{arrow}{abs(c)*100:.1f}%</span>')
        d['cat_changes_text'] = ' | '.join(changes) if changes else '—'
        print(f"  Categories: {len(cats_with_data)} cats, months: {[month_names_he[m] for y,m in months_back]}")
    except Exception as e:
        print(f'  Category computation failed: {e}')
        d['cats_json']  = ''
        d['cat_months'] = "'מרץ','אפריל','מאי'"
        d['cat_range']  = 'מרץ עד מאי' 
    # Holdings
    d['holdings'] = []
    try:
        ws_inv = wb['תיק השקעות עדכני']
        for row in ws_inv.iter_rows(values_only=True):
            if row[0] and str(row[0]).isdigit() and row[1] and row[5]:
                d['holdings'].append({
                    'name': str(row[1]), 'value': float(row[5]),
                    'pct': float(row[9] or 0), 'chg': float(row[7] or 0)
                })
    except: pass

    wb.close()

    # Derived
    d['pension_all']       = d['pension_total'] + d['hishtalmut_total']
    d['rsu_total']         = d['rsu_vested'] + d['rsu_unvested']
    d['income_children'] = 0  # קצבת ילדים — not in workbook currently
    d['income_other']    = 0
    d['income_other_note'] = ''
    d['exp_other']       = max(0, d.get('expenses', 0) - abs(d.get('exp_credit', 0)) - d.get('mortgage', 0))
    exp_total_for_other  = d.get('expenses', 0) or 1
    d['exp_other_pct']   = f"{d['exp_other']/exp_total_for_other*100:.1f}"

    try:
        ws_hist = xw_wb.sheets['היסטוריה']
        rate_val = ws_hist.range((25, 14)).value
        d['rate'] = float(rate_val) if rate_val and isinstance(rate_val, (int,float)) else 3.65
    except:
        d['rate'] = 3.65
    d['rsu_ils']           = d['rsu_total'] * d['rate']
    d['rsu_vested_ils']    = d['rsu_vested'] * d['rate']
    d['liquid']            = d['portfolio_val'] + d['bank'] + d['rsu_ils']
    d['gross']             = d['liquid'] + d['pension_all']
    # net_worth read directly from K32; gross computed for display
    d['kaspiyot_excess']   = max(0, d['kaspiyot'] - 350000)
    d['invest_pct']        = int(d['portfolio_val'] / 2000000 * 100)
    d['stocks_pct']        = 71
    d['mm_pct']            = 29
    # Approx ALGN price from RSU vested / vested shares
    # Compute ALGN price from vested value / vested shares (1,000 total vested)
    vested_shares = 76 + 99 + 245 + 362 + 218  # total vested as of May 2026
    d['algn_price'] = round(d['rsu_vested'] / vested_shares, 2) if vested_shares > 0 else 163.38

    return d


def build_top5_html(top5):
    rows = ''
    for e in top5:
        rows += (f'      <div class="exp-row">'
                 f'<span class="exp-row-name">{e["name"]}</span>'
                 f'<span class="exp-row-badge">{e["source"]}</span>'
                 f'<span class="exp-row-amount">{ils(e["amount"])}</span></div>\n')
    return rows

def build_top3_html(top3):
    if not top3: return ''
    max_amt = max(c['amount'] for c in top3)
    rows = ''
    for c in top3:
        bar_w = int(c['amount'] / max_amt * 100) if max_amt else 0
        pct_chg = c['pct'] - c['prev_pct']
        arrow = '▲' if pct_chg >= 0 else '▼'
        clr = 'var(--red)' if pct_chg >= 0 else 'var(--green)'
        rows += (f'      <div class="cat-row">\n'
                 f'        <div class="cat-name">{c["name"]}</div>\n'
                 f'        <div class="cat-bar-wrap"><div class="cat-bar-fill" style="width:{bar_w}%"></div></div>\n'
                 f'        <div class="cat-amount">{ils(c["amount"])}</div>\n'
                 f'        <div class="cat-pct">{pct(c["pct"])}</div>\n'
                 f'      </div>\n')
    return rows

def build_holdings_html(holdings):
    rows = ''
    for h in holdings:
        arrow = '▲' if h['chg'] >= 0 else '▼'
        clr = 'c-green' if h['chg'] >= 0 else 'c-red'
        rows += (f'            <tr>\n'
                 f'              <td>{h["name"]}</td>\n'
                 f'              <td style="font-family:var(--mono)">{ils(h["value"])}</td>\n'
                 f'              <td style="font-family:var(--mono);color:var(--muted)">{pct(h["pct"])}%</td>\n'
                 f'              <td style="font-family:var(--mono);color:var(--{clr.replace("c-","")})">{arrow} {abs(h["chg"]*100):.1f}%</td>\n'
                 f'            </tr>\n')
    return rows

_MONTH_MAP = {
    'January':'ינו','February':'פבר','March':'מרץ','April':'אפר',
    'May':'מאי','June':'יוני','July':'יולי','August':'אוג',
    'September':'ספט','October':'אוק','November':'נוב','December':'דצמ'
}
def _month_he(period):
    for eng, heb in _MONTH_MAP.items():
        if eng in period: return heb
    return period[:3]

def fill_template(template, d):
    # Build dynamic HTML blocks
    top5_html     = build_top5_html(d['top5'])
    top3_html     = build_top3_html(d['top3'])
    holdings_html = build_holdings_html(d['holdings'])

    # Kaspiyot deploy note
    monthly_chunk = d['kaspiyot_excess'] / 3
    deploy_note = f'פרוס: {ils(monthly_chunk)} × 3 חודשים לתוך ACWI.' if d['kaspiyot_excess'] > 0 else 'הכספית בטווח היעד — אין צורך בפעולה.'

    # RSU pct from target
    target = 300
    rsu_pct_from_target = f'+{int((target - d["algn_price"]) / d["algn_price"] * 100)}% ליעד'

    substitutions = {
        '__PERIOD__':             str(d['period']),
        '__CURRENT_MONTH__':       _month_he(str(d['period'])),
        '__REPORT_DATE__':        str(d['report_date']),
        '__LIQUID__':             ils(d['liquid']),
        '__PENSION_ALL__':        ils(d['pension_all']),
        '__NET_WORTH__':          ils(d['net_worth']),
        '__MORTGAGE_NEG__':       '−' + ils(d['mortgage']),
        '__GROSS__':              ils(d['gross']),
        '__RATE__':               f'₪{d["rate"]:.4f}',
        '__RATE_DATE__':          str(d['report_date']),
        '__INCOME__':             ils(d['income']),
        '__EXPENSES__':           ils(d['expenses']),
        '__SURPLUS__':            ils(d['surplus']),
        '__SAVINGS_RATE__':       pct(d['savings_rate']),
        '__BANK_FREE_STRIP__':    ils(d['invest_surplus']),
        '__BANK_FREE__':          ils(d['bank_free']),
        '__MORTGAGE__':           ils(d['mortgage']),
        '__INCOME_ALIGN__':       ils(d['income']),
        '__INCOME_OTHER__':       '—',
        '__EXP_CREDIT__':         ils(d['exp_credit']),
        '__AVG_MONTHLY__':        ils(d['expenses']),
        '__BANK__':               ils(d['bank']),
        '__SEC_VAL__':            ils(d['portfolio_val'] - d['kaspiyot']),
        '__CREDIT_NEG__':         '—',
        '__PORTFOLIO__':          ils(d['portfolio_val']),
        '__INVEST_PCT__':         str(d['invest_pct']),
        '__CUMUL_GAIN__':         ils(d['cumulative_gain']),
        '__RETURN_PCT__':         pct(d['return_pct']),
        '__MONTHLY_RET__':        pct(d['monthly_return']),
        '__RSU_USD__':            usd(d['rsu_total']),
        '__RSU_ILS__':            ils(d['rsu_ils']),
        '__ALGN_PRICE__':         f'${d["algn_price"]:,.2f}',
        '__ALGN_PRICE_RAW__':     str(d['algn_price']),
        '__ALGN_FROM_HIGH__':     f'−{int((1 - d["algn_price"]/735)*100)}% מהשיא',
        '__ALGN_TO_TARGET__':     f'+{int((300/d["algn_price"]-1)*100)}%',
        '__STOCKS_PCT__':         str(d.get('stocks_pct', 71)),
        '__MM_PCT__':             str(d.get('mm_pct', 29)),
        '__RSU_VESTED_USD__':     usd(d['rsu_vested']),
        '__RSU_VESTED_ILS__':     ils(d['rsu_vested_ils']),
        '__PENSION_TOTAL_K__':    k(d['pension_total']),
        '__HISHTALMUT_TOTAL_K__': k(d['hishtalmut_total']),
        '__PENSION_CHG__':        ils_signed(d['pension_dror_chg']),
        '__PENSION_DROR_F1__':    ils(d.get('pension_dror_f1', d['pension_dror'] * 0.5)),
        '__PENSION_DROR_F2__':    ils(d.get('pension_dror_f2', d['pension_dror'] * 0.5)),
        '__PENSION_DROR__':       ils(d['pension_dror']),
        '__PENSION_LIAT__':       ils(d['pension_liat']),
        '__HISHTALMUT_DROR_F1__': ils(d['hishtalmut_dror_f1']),
        '__HISHTALMUT_DROR_F2__': ils(d['hishtalmut_dror_f2']),
        '__HISHTALMUT_LIAT__':    ils(d['hishtalmut_liat']),
        '__HISHTALMUT_TOTAL__':   ils(d['hishtalmut_total']),
        '__PENSION_MONTHLY__':    ils(d['pension_monthly']),
        '__INVEST_SURPLUS__':     ils(d['invest_surplus']),
        '__TARGET_DIST__':        ils(d['target_dist']),
        '__STOCKS_PCT__':         str(d['stocks_pct']),
        '__MM_PCT__':             str(d['mm_pct']),
        '__KASPIYOT__':           ils(d['kaspiyot']),
        '__KASPIYOT_EXCESS__':    ils(d['kaspiyot_excess']),
        '__KASPIYOT_DEPLOY_NOTE__': deploy_note,
        '__RSU_PCT_FROM_TARGET__': rsu_pct_from_target,
        '__TOP5_ROWS__':          top5_html,
        '__CHART_MONTHS__':        ','.join(f"'{m}'" for m in d.get('chart_months', [])),
        '__CHART_INCOME__':        ','.join(str(v) for v in d.get('chart_income', [])),
        '__CHART_EXPENSES__':      ','.join(str(v) for v in d.get('chart_expenses', [])),
        '__CHART_INVEST__':        ','.join(str(v) for v in d.get('chart_invest', [])),
        '__CHART_SURPLUS__':       ','.join(str(v) for v in d.get('chart_surplus', [])),
        '__CHART_ALIGN__':         ','.join(str(v) for v in d.get('chart_align', [])),
        '__EXP_CREDIT_PCT__':      d.get('exp_credit_pct', '0'),
        '__EXP_CASH__':            d.get('exp_cash', '₪0'),
        '__EXP_CASH_PCT__':        d.get('exp_cash_pct', '0'),
        '__EXP_MORTGAGE__':        d.get('exp_mortgage', '₪0'),
        '__EXP_MORTGAGE_PCT__':    d.get('exp_mortgage_pct', '0'),
        '__CATS_JSON__':           d.get('cats_json', ''),
        '__CAT_MONTHS__':          d.get('cat_months', "'ינואר','פברואר','מרץ'"),
        '__CAT_RANGE__':           d.get('cat_range', 'ינואר עד מרץ'),
        '__CAT_RANGE_START__':     d.get('cat_range_start', 'ינואר'),
        '__CAT_TOTAL_CHG__':       d.get('cat_total_chg', '—'),
        '__CAT_TOTAL_CHG_COLOR__': d.get('cat_total_chg_color', 'c-muted'),
        '__CAT_CHANGES_TEXT__':    d.get('cat_changes_text', '—'),
        '__INCOME_CHILDREN__':     ils(d.get('income_children', 0)),
        '__INCOME_OTHER__':        ils(d.get('income_other', 0)),
        '__INCOME_OTHER_NOTE__':   d.get('income_other_note', ''),
        '__EXP_OTHER__':           ils(d.get('exp_other', 0)),
        '__EXP_OTHER_PCT__':       d.get('exp_other_pct', '0'),
        '__CHART_INSIGHTECH__':    ','.join(str(v) for v in d.get('chart_insightech', [])),
        '__TOP3_ROWS__':          top3_html,
        '__HOLDINGS_ROWS__':      holdings_html,
    }

    for placeholder, value in substitutions.items():
        template = template.replace(placeholder, str(value))

    return template


def upload_to_fileio(html_path):
    import base64 as b64
    with open(html_path, 'r', encoding='utf-8') as f:
        content = f.read()
    gh_token = open(BASE / 'gh_token.txt').read().strip()
    # Push dashboard.html to GitHub repo, serve via htmlpreview.github.io
    api_url = 'https://api.github.com/repos/drorklein-boop/klein-finance/contents/dashboard_live.html'
    # Get current SHA if file exists
    file_sha = None
    try:
        r = urllib.request.urlopen(urllib.request.Request(api_url,
            headers={'Authorization': 'token ' + gh_token, 'User-Agent': 'KleinFinance'}))
        file_sha = json.loads(r.read()).get('sha')
    except: pass
    encoded = b64.b64encode(content.encode('utf-8')).decode('ascii')
    payload = json.dumps({'message': 'update dashboard', 'content': encoded,
                          **({'sha': file_sha} if file_sha else {})}).encode('utf-8')
    try:
        req = urllib.request.Request(api_url, data=payload, method='PUT',
            headers={'Authorization': 'token ' + gh_token,
                     'Content-Type': 'application/json', 'User-Agent': 'KleinFinance'})
        with urllib.request.urlopen(req, timeout=30) as r:
            json.loads(r.read())
        return 'https://htmlpreview.github.io/?https://raw.githubusercontent.com/drorklein-boop/klein-finance/main/dashboard_live.html'
    except Exception as e:
        print(f'  Upload failed: {e}')
    return None

def main():
    print('\n  Klein Finance - Dashboard Generator v3.1')
    print('  =========================================')

    data = read_workbook()
    if not data:
        input('\n  Press Enter to close...'); return

    print('  Fetching template...')
    template = get_template()

    print('  Filling data...')
    html = fill_template(template, data)

    out = BASE / 'dashboard.html'
    out.write_text(html, encoding='utf-8')
    print(f'  Dashboard saved: {out}')

    webbrowser.open(out.as_uri())
    print('  Opened in browser.')

    print('  Uploading for sharing...')
    link = upload_to_fileio(out)
    if link:
        print('\n  =====================================')
        print('  SHAREABLE LINK:')
        print(f'  {link}')
        print('  =====================================')
    else:
        print('  Upload failed — dashboard available locally only.')

    input('\n  Press Enter to close...')

if __name__ == '__main__':
    main()
