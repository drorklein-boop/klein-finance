# Klein Finance - Monthly Sheet Updater v8.0
import sys, shutil, datetime, json, base64, re
from pathlib import Path
from collections import defaultdict

BASE    = Path(__file__).parent
MONTHLY = BASE / "MONTHLY"
TRACKER = BASE / "processed_files.json"
API_KEY_FILE = BASE / "api_key.txt"

DROR_POLICY = '35995836'
LIAT_POLICY = '6650891010'

def load_tracker():
    if TRACKER.exists():
        try: return json.loads(TRACKER.read_text(encoding='utf-8'))
        except: pass
    return {}

def save_tracker(tracker):
    TRACKER.write_text(json.dumps(tracker, indent=2, ensure_ascii=False), encoding='utf-8')

def file_sig(fpath):
    s = fpath.stat()
    return {"mtime": round(s.st_mtime, 2), "size": s.st_size}

def is_new(fpath, tracker):
    key = fpath.name
    sig = file_sig(fpath)
    if key not in tracker: return True
    t = tracker[key]
    return t.get('mtime') != sig['mtime'] or t.get('size') != sig['size']

def detect_by_name(fname):
    if fname.startswith('~'): return None
    name = fname.lower()
    if name.endswith(('.jpeg','.jpg','.png')): return 'rsu_image'
    if 'transaction-details' in name: return 'credit'
    if 'עוש' in name or 'לאומי' in name: return 'bank'
    if 'אחזקות' in name: return 'invest'
    if 'התמונה המלאה' in name: return 'pension_check'
    if '5647' in name or 'אישראכרט' in name: return 'isracard'
    if 'ריכוז' in name and 'יתרות' in name: return 'balance'
    return None

def detect_by_content(fpath):
    fname = fpath.name.lower()
    is_xlsx = fname.endswith('.xlsx') or fname.endswith('.xls.xlsx')
    is_xls  = fname.endswith('.xls') and not fname.endswith('.xls.xlsx')
    try:
        if is_xlsx:
            import openpyxl
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            sheets = set(wb.sheetnames)
            first_rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))[:3]
            wb.close()
            if 'עסקאות במועד החיוב' in sheets or 'עסקאות חו״ל ומט״ח' in sheets: return 'credit'
            if 'פירוט עסקאות' in sheets: return 'isracard'
            if 'עוש' in sheets: return 'bank'
            for row in first_rows:
                if any('מבט אישי' in str(v or '') for v in row): return 'invest'
        elif is_xls:
            import xlrd
            wb = xlrd.open_workbook(fpath)
            if 'פרטי המוצרים שלי' in set(wb.sheet_names()):
                ws = wb.sheet_by_name('פרטי המוצרים שלי')
                vals = ' '.join(str(ws.cell(r,c).value) for r in range(ws.nrows) for c in range(ws.ncols))
                if DROR_POLICY in vals: return 'pension_dror'
                if LIAT_POLICY in vals: return 'pension_liat'
    except: pass
    return None

def detect(fpath):
    by_name = detect_by_name(fpath.name)
    if by_name == 'pension_check' or by_name is None:
        return detect_by_content(fpath)
    return by_name

def read_rsu_from_image(img_path):
    try:
        import anthropic
    except ImportError:
        import subprocess
        subprocess.run([sys.executable,'-m','pip','install','anthropic','--quiet'], check=True)
        import anthropic

    api_key = API_KEY_FILE.read_text().strip()
    img_b64 = base64.standard_b64encode(img_path.read_bytes()).decode()
    ext = img_path.suffix.lower().lstrip('.')
    media_type = 'image/jpeg' if ext in ('jpg','jpeg') else f'image/{ext}'

    client = anthropic.Anthropic(api_key=api_key)
    msg = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=256,
        messages=[{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": img_b64}},
                {"type": "text", "text": (
                    "This is an RSU portfolio screenshot. Find two dollar amounts:\n"
                    "1. available/vested (zamin lemimoosh)\n"
                    "2. unvested (trem hivshil)\n"
                    'Return ONLY this JSON with no markdown, no explanation: {"available": 170600.00, "unvested": 187148.20}'
                )}
            ]
        }]
    )
    raw = msg.content[0].text.strip()
    raw = re.sub(r'^```[a-z]*\s*', '', raw)
    raw = re.sub(r'\s*```$', '', raw)
    result = json.loads(raw.strip())
    return float(result['available']), float(result['unvested'])

def clean_val(val):
    if isinstance(val, str):
        for ch in ('\u200e','\u200f','\u202b','\u202c'): val = val.replace(ch,'')
        val = val.strip()
        try: return float(val.replace(',',''))
        except: return val if val else None
    return val

def read_full_xlsx(path, sheet_name=None):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name or wb.sheetnames[0]]
    data = [[clean_val(v) for v in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    return data

def read_full_xls(path, sheet_name):
    import xlrd
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_name(sheet_name)
    return [[clean_val(ws.cell(r,c).value) for c in range(ws.ncols)] for r in range(ws.nrows)]

def read_from_header(path, header_val, sheet_name=None):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name or wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    start = next((i for i,r in enumerate(rows) if r[0] and str(r[0]).strip()==header_val), None)
    if start is None: raise ValueError(f"Header not found")
    return [[clean_val(v) for v in row] for row in rows[start:]]

def read_file(ftype, fpath):
    fname = fpath.name.lower()
    is_xlsx = fname.endswith('.xlsx') or fname.endswith('.xls.xlsx')
    is_xls  = fname.endswith('.xls') and not fname.endswith('.xls.xlsx')

    if ftype == 'credit' and is_xlsx:
        import openpyxl
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        out = {s: [[clean_val(v) for v in row] for row in wb[s].iter_rows(values_only=True)] for s in wb.sheetnames}
        wb.close()
        return out
    elif ftype == 'bank' and is_xlsx:
        return {'עוש': read_full_xlsx(fpath, 'עוש')}
    elif ftype == 'invest' and is_xlsx:
        import openpyxl
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        for s in wb.sheetnames:
            first = next(wb[s].iter_rows(values_only=True), [])
            if first and 'מבט אישי' in str(first[0] or ''):
                data = [[clean_val(v) for v in row] for row in wb[s].iter_rows(values_only=True)]
                wb.close()
                return {'תיק השקעות עדכני': data}
        wb.close()
        return {}
    elif ftype == 'pension_dror' and is_xls:
        return {'דרור - מסלקה': read_full_xls(fpath, 'פרטי המוצרים שלי')}
    elif ftype == 'pension_liat' and is_xls:
        return {'ליאת - מסלקה': read_full_xls(fpath, 'פרטי המוצרים שלי')}
    elif ftype == 'isracard' and is_xlsx:
        import pandas as pd
        isr_files = sorted(
            [f for f in MONTHLY.glob('*.xls*') if '5647' in f.name or 'אישראכרט' in f.name],
            key=lambda f: f.stat().st_mtime
        )
        frames = []
        for ipath in isr_files:
            rows = read_from_header(ipath, 'תאריך רכישה')
            if len(rows) < 2:
                continue
            df = pd.DataFrame(rows[1:], columns=rows[0])
            frames.append(df)
        if not frames:
            return {}
        combined = pd.concat(frames, ignore_index=True)
        key_cols = [col for col in ["מס' שובר", 'תאריך רכישה', 'סכום עסקה'] if col in combined.columns]
        if key_cols:
            combined = combined.drop_duplicates(subset=key_cols)
        if 'תאריך רכישה' in combined.columns:
            combined['תאריך רכישה'] = pd.to_datetime(combined['תאריך רכישה'], dayfirst=True, errors='coerce')
            combined = combined.sort_values('תאריך רכישה', ascending=True)
        sheet_cols = ['תאריך רכישה', 'שם בית עסק', 'סכום עסקה', 'מטבע עסקה', 'סכום חיוב', 'מטבע חיוב', "מס' שובר", 'פירוט נוסף', 'קטגוריה']
        for col in sheet_cols:
            if col not in combined.columns:
                combined[col] = None
        combined = combined[sheet_cols]
        return {'אישראכרט': combined.values.tolist()}
    elif ftype == 'balance' and is_xlsx:
        return {'ריכוז יתרות לאומי': read_from_header(fpath, 'סוג פעילות')}
    return {}

def write_sheet(xw_wb, name, data):
    ws = xw_wb.sheets[name]
    if name == 'אישראכרט':
        ws.range('A2:Z10000').clear_contents()
        if data:
            xw_wb.app.screen_updating = False
            xw_wb.app.calculation = 'manual'
            ws.range('A2').value = data
            xw_wb.app.calculation = 'automatic'
            xw_wb.app.screen_updating = True
    else:
        ws.clear_contents()
        if data:
            xw_wb.app.screen_updating = False
            xw_wb.app.calculation = 'manual'
            ws.range('A1').value = data
            xw_wb.app.calculation = 'automatic'
            xw_wb.app.screen_updating = True

def main():
    print("\n  Klein Finance - Monthly Update v8.0")
    print("  =====================================")

    try:
        import xlwings as xw
    except ImportError:
        import subprocess
        subprocess.run([sys.executable,'-m','pip','install','xlwings','--quiet'], check=True)
        import xlwings as xw
    try:
        import openpyxl, xlrd
    except ImportError:
        import subprocess
        subprocess.run([sys.executable,'-m','pip','install','openpyxl','xlrd','--quiet'], check=True)

    app = xw.apps.active
    if not app:
        print("  ERROR: Excel is not open.")
        input("\n  Press Enter to close..."); return

    # Find workbook by .xlsm extension — avoids Hebrew encoding issues
    wb = next((b for b in app.books if b.name.lower().endswith('.xlsm')), None)
    if not wb:
        print("  ERROR: No .xlsm workbook open in Excel.")
        input("\n  Press Enter to close..."); return

    print(f"  Workbook found: {wb.name}")

    backup_dir = BASE / "backups"
    backup_dir.mkdir(exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y-%m-%d_%H%M")
    src = BASE / wb.name
    if src.exists():
        shutil.copy2(src, backup_dir / f"backup_{ts}.xlsm")
        print("  Backup saved")

    tracker = load_tracker()
    all_files = [f for f in MONTHLY.iterdir()
                 if f.is_file() and not f.name.startswith('~')
                 and f.suffix.lower() not in ('.xml',)]
    new_files = [f for f in all_files if is_new(f, tracker)]

    if not new_files:
        print("\n  No new or changed files since last run.")
        input("\n  Press Enter to close..."); return

    print(f"\n  New/changed files: {len(new_files)}")

    typed = defaultdict(list)
    for f in new_files:
        t = detect(f)
        if t: typed[t].append(f)
    for t in typed:
        typed[t].sort(key=lambda f: f.stat().st_mtime, reverse=True)

    if not typed:
        print("  None identified.")
        input("\n  Press Enter to close..."); return

    print("  Detected:")
    for ftype, files in typed.items():
        print(f"    {ftype}: {files[0].name}")

    results = []
    target_sheets = [s.name for s in wb.sheets]
    successfully_processed = set()

    for ftype, files in typed.items():
        fpath = files[0]
        if ftype == 'rsu_image':
            try:
                available, unvested = read_rsu_from_image(fpath)
                wb.sheets['ALIGN RSU']['H13'].value = available
                wb.sheets['ALIGN RSU']['H14'].value = unvested
                results.append(f"  OK    ALIGN RSU  H13={available:,.2f}  H14={unvested:,.2f}")
                successfully_processed.add(fpath.name)
            except Exception as e:
                results.append(f"  ERROR rsu ({fpath.name}): {e}")
            continue
        try:
            sheets_data = read_file(ftype, fpath)
            for tname, data in sheets_data.items():
                if tname not in target_sheets:
                    results.append(f"  SKIP  '{tname}' not in workbook"); continue
                write_sheet(wb, tname, data)
                results.append(f"  OK    {tname} ({len(data)} rows)")
            successfully_processed.add(fpath.name)
        except Exception as e:
            results.append(f"  ERROR {ftype} ({fpath.name}): {e}")

    wb.save()

    updated_tracker = dict(tracker)
    for f in all_files:
        is_queued = any(f.name == ff.name for files in typed.values() for ff in files)
        if not is_queued or f.name in successfully_processed:
            updated_tracker[f.name] = file_sig(f)
    save_tracker(updated_tracker)

    print("\n  Results:")
    for r in results: print(r)
    print("\n  Done.")
    input("\n  Press Enter to close...")

if __name__ == '__main__':
    main()
